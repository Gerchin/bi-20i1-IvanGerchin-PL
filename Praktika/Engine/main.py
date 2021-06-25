from openpyxl import load_workbook
from datetime import datetime
from data import *
from matplotlib.pyplot import *

workbook = load_workbook('vullist.xlsx')
sheet = workbook.active

#print(datetime.strptime(td, formatter))

counter = 0

'''
colum = {
	'2': 'Наименование уязвимости',
	'3': 'Описание уязвимости',
	'5': 'Название ПО',
	'7': 'Тип ПО',
	'9': 'Класс уязвимости',
	'10': 'Дата выявления',
	'13': 'Уровень опасности уязвимости',
	'15': 'Статус уязвимости',
	'16': 'Наличие эксплойта',
	'17': 'Информация об устранении',
	'21': 'Описание ошибки CWE'}
'''

for row in range(1, sheet.max_row):
	if 'Microsoft Edge' in str(sheet.cell(row=row, column=5).value):
		nameU.append(sheet.cell(row=row, column=2).value)			#0
		discription.append(sheet.cell(row=row, column=3).value)		#1
		soft.append(sheet.cell(row=row, column=5).value)			#2
		softType.append(sheet.cell(row=row, column=7).value)		#3
		classU.append(sheet.cell(row=row, column=9).value)			#4
		dateOfFind.append(sheet.cell(row=row, column=10).value)		#5
		dangerLevel.append(sheet.cell(row=row, column=13).value)	#6
		status.append(sheet.cell(row=row, column=15).value)			#7
		explExists.append(sheet.cell(row=row, column=16).value)		#8
		info.append(sheet.cell(row=row, column=17).value)			#9
		cwe.append(sheet.cell(row=row, column=21).value)			#10
		#print(soft[counter], nameU[counter], dateOfFind[counter])
		counter +=1

for i in range(counter):
		allElements[i] = (nameU[i], discription[i], soft[i], softType[i], classU[i], dateOfFind[i], dangerLevel[i], status[i], explExists[i], info[i], cwe[i])
		#print(allElements[i])

def countCheck(listToCount):
	ct = 0
	for t in listToCount:
		ct += 1
	return ct

def lookFor(look, listToCount):
	ct = 0
	for p in listToCount:
		for x in allElements.get(p):
			if str(look) in x:
				print(x)
				ct += 1
	return ct

print(countCheck(allElements))
print('Критический',lookFor('Критический', allElements))
#print(allElements.get(696, 5))

def date_perform(date):
	formatter = "%d.%m.%Y"
	formatterForQtDate = "%Y-%m-%d"
	dateP = datetime.strptime(date, formatter)
	return dateP

def sort(date_until, date_after, dangerLevel=0, status=0, explExists=0):
	for i in range(countCheck(allElements)):
		if (date_perform(allElements[i][5]) > date_perform(date_after)) | (date_perform(allElements[i][5]) < date_perform(date_until)):
			del allElements[i]
			continue

		if dangerLevel != 0:
			if (dangerLevel != 1) & ('Низкий' in allElements[i][6]):
				del allElements[i]
				continue

			elif (dangerLevel != 2) & ('Средний' in allElements[i][6]):
				del allElements[i]
				continue

			elif (dangerLevel != 3) & ('Высокий' in allElements[i][6]):
				del allElements[i]
				continue

			elif (dangerLevel != 4) & ('Критический' in allElements[i][6]):
				del allElements[i]
				continue

		if status != 0:
			if (status != 1) & ('Подтверждена' in allElements[i][7]):
				del allElements[i]
				continue

			elif (status != 2) & ('Потенциальная' in allElements[i][7]):
				del allElements[i]
				continue

		if explExists != 0:
			if (explExists != 1) & ('Существует' in allElements[i][8]):
				del allElements[i]
				continue

			elif (explExists != 2) & ('уточняются' in allElements[i][8]):
				del allElements[i]
				continue

	#sorted_data = [[0]*11 for i in range(counter)] sozdanie dvumernogo massiva

print(countCheck(allElements))
#print(allElements.keys())

sort('12.03.2013', '21.05.2022', dangerLevel=4)
print(countCheck(allElements))
#print(allElements)
k = 0

def printSort():
	for i in allElements:
		print('Наименование уязвимости:', allElements[i][k])
		print('Описание уязвимости:', allElements[i][k+1])
		print('Название ПО:', allElements[i][k+2])
		print('Тип ПО:', allElements[i][k+3])
		print('Класс уязвимости:', allElements[i][k+4])
		print('Дата выявления:', allElements[i][k+5])
		print('Уровень опасности уязвимости:', allElements[i][k+6])
		print('Статус уязвимости:', allElements[i][k+7])
		print('Наличие эксплойта:', allElements[i][k+8])
		print('Информация об устранении:', allElements[i][k+9])
		print('')

#plot(nameU)
#show()